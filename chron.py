import io
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
import re

def convert_string_data(value):
    """
    Convert a string to the appropriate type (int, float, or string)
    
    Args:
        value: The string value to convert

    Returns:
        The converted value as an int, float, or string
    """
    
    if value is None:
        return ""
                
    # Clean the string
    clean_value = value.strip()
    
    # Check for empty string
    if not clean_value or clean_value == "N/A":
        return ""
        
    # Try to convert to int
    try:
        # Check if it's a simple integer (no decimal point)
        if clean_value.isdigit() or (clean_value.startswith('-') and clean_value[1:].isdigit()):
            return int(clean_value)
    except (ValueError, TypeError):
        pass
        
    # Try to convert to float (numbers with decimal points)
    try:
        # This regex matches numbers like 1,234.56 and converts them to 1234.56
        if re.match(r'^-?\d{1,3}(,\d{3})*(\.\d+)?$', clean_value):
            clean_value = clean_value.replace(',', '')
        if '.' in clean_value:
            return float(clean_value)
    except (ValueError, TypeError):
        pass
        
    # If all conversions fail, return the original string
    return clean_value

def format_excel_table(worksheet, data, title="Returns Data Chronology", footnotes=None, include_col_numbers=True):
    """
    Format the Excel worksheet with styling and apply footnotes

    Args:
        worksheet: openpyxl worksheet object
        data: 2D array of data (first row is headers)
        title: Title for the Excel file
        footnotes: Dictionary of footnotes
        include_col_numbers: Whether to include column numbers in the output

    Returns:
        Formatted worksheet
    """


    # Process footnotes dictionary - convert from element_id format to simpler format
    processed_footnotes = {}
    if footnotes:
       
        # Iterate over the footnotes dictionary.
        # Each key (element_id) is formatted as either "header_<col>" for header footnotes 
        # or "cell_<row>_<col>" for cell footnotes, with the value being the corresponding text.
        for element_id, text in footnotes.items():
            if not text:  # Skip empty footnotes
                continue

            parts = element_id.split('_')  # Example: header_1 or cell_1_2
            if len(parts) < 3:
                print(f"Invalid footnote element ID: {element_id}")
                continue

            header_or_cell = parts[1]  # header or cell

            if header_or_cell == 'header':
                if len(parts) >= 3:
                    col_idx = int(parts[2])
                    processed_footnotes[f"header_{col_idx}"] = text
            elif header_or_cell == 'cell' and len(parts) >= 4:
                row_idx = int(parts[2])
                col_idx = int(parts[3])
                processed_footnotes[f"cell_{row_idx}_{col_idx}"] = text

    # Create superscript style with Times New Roman
    superscript = NamedStyle(name="superscript")
    superscript.font = Font(vertAlign="superscript", size=7, name="Times New Roman")
    worksheet.parent.add_named_style(superscript)

    # Detect significance indicator columns (columns containing only "*", "**", "***", "N/A" or empty)
    significance_cols = []
    for col_idx in range(len(data[0])):
        is_significance_col = True
        for row_idx in range(1, len(data)):  # Skip header row
            cell_value = str(data[row_idx][col_idx]).strip() if data[row_idx][col_idx] is not None else ""
            if cell_value not in ["*", "**", "***", "N/A", ""]:
                is_significance_col = False
                break
        if is_significance_col and col_idx > 0:  # Must have a column to the left
            significance_cols.append(col_idx)

    # Get number of original columns and calculate total columns with spacers
    # Exclude significance columns from the count as they'll be merged
    orig_cols = len(data[0]) - len(significance_cols) if data and len(data) > 0 else 0
    total_cols = orig_cols * 2 - 1 if orig_cols > 0 else 1

    # Add title row, merged across all columns, with Times New Roman
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    worksheet.cell(row=1, column=1, value=title)
    worksheet.cell(1, 1).font = Font(size=14, bold=True, name="Times New Roman")
    worksheet.cell(1, 1).alignment = Alignment(horizontal='center')

    # Define styles with Times New Roman
    header_font = Font(bold=True, name="Times New Roman")
    bottom_border = Border(bottom=Side(style='medium'))

    # Starting from row 3, leaving row 2 empty
    current_row = 3

    # Track all footnotes to add at the bottom of the document
    all_footnotes = {}
    footnote_counter = 1

    # Add headers with spacing, skipping significance columns
    if data and len(data) > 0:
        headers = data[0]
        output_col = 1  # Track the actual Excel column position

        for i, header in enumerate(headers):
            if i in significance_cols:
                continue

            # Calculate actual column (with spacers)
            col_idx = output_col
            output_col += 2  # Increment for next column and spacer

            # Check if this header has a footnote
            header_id = f"header_{i}"
            if header_id in processed_footnotes and processed_footnotes[header_id]:
                worksheet.cell(row=current_row, column=col_idx, value=header)
                footnote_cell = worksheet.cell(row=current_row, column=col_idx+1, value=str(footnote_counter))
                footnote_cell.style = "superscript"

                # Track the footnote
                all_footnotes[footnote_counter] = processed_footnotes[header_id]
                footnote_counter += 1
            else:
                worksheet.cell(row=current_row, column=col_idx, value=header)

            cell = worksheet.cell(row=current_row, column=col_idx)
            cell.font = header_font
            cell.border = bottom_border

            # Make column wider
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = max(len(str(header)) + 2, 12)

            # Set width for spacer column
            spacer_col = get_column_letter(col_idx + 1)
            worksheet.column_dimensions[spacer_col].width = 3

    # Add column numbers if requested
    if include_col_numbers:
        current_row += 1
        output_col = 1
        for i, header in enumerate(headers):
            if i in significance_cols:
                continue
            col_idx = output_col
            output_col += 2
            cell = worksheet.cell(row=current_row, column=col_idx, value=f"({i+1})")
            cell.font = Font(bold=True, size=9, name="Times New Roman")
            cell.alignment = Alignment(horizontal='center')

        current_row += 1
    else:
        current_row += 1

    # Add data rows
    if data and len(data) > 1:
        for row_idx, row_data in enumerate(data[1:]):
            output_col = 1
            last_main_output_col = None
            for i, cell_value in enumerate(row_data):
                
                # Convert the cell value to the appropriate type
                cell_value = convert_string_data(cell_value)
                
                # If this is a significance column, place indicator in the corresponding spacer column
                if i in significance_cols:
                    if last_main_output_col is not None and cell_value:
                        worksheet.cell(row=current_row, column=last_main_output_col+1, value=cell_value)
                    continue

                col_idx = output_col
                last_main_output_col = col_idx
                output_col += 2

                cell_id = f"cell_{row_idx}_{i}"
                if cell_id in processed_footnotes and processed_footnotes[cell_id]:
                    # Set cell with properly typed value
                    cell = worksheet.cell(row=current_row, column=col_idx, value=cell_value)
                    # Set font
                    cell.font = Font(name="Times New Roman")
                    
                    footnote_cell = worksheet.cell(row=current_row, column=col_idx+1, value=f"{footnote_counter}")
                    footnote_cell.style = "superscript"
                    all_footnotes[footnote_counter] = processed_footnotes[cell_id]
                    footnote_counter += 1
                else:
                    # Setting properly typed value with Times New Roman
                    cell = worksheet.cell(row=current_row, column=col_idx, value=cell_value)
                    cell.font = Font(name="Times New Roman")
            current_row += 1

    # Add footnotes at the bottom
    if all_footnotes:
        current_row += 2  # Add some space
        footnote_heading = worksheet.cell(row=current_row, column=1, value="Footnotes:")
        footnote_heading.font = Font(bold=True, name="Times New Roman")
        current_row += 1

        for num, text in all_footnotes.items():
            note_cell = worksheet.cell(row=current_row, column=1, value=f"{num}. {text}")
            note_cell.font = Font(name="Times New Roman")
            worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=min(5, total_cols))
            note_cell.alignment = Alignment(wrap_text=True)
            current_row += 1

    return worksheet

def create_excel_from_table_data(table_data, title="Returns Data Chronology", footnotes=None):
    """
    Create a styled Excel file from table data
    
    Args:
        table_data: 2D array of data (first row is headers)
        title: Title for the Excel file
        footnotes: Dictionary of footnotes
        
    Returns:
        BytesIO object containing the Excel file
    """
    # Create workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Chronology"
    
    if footnotes is None:
        footnotes = {}
    
    # Format the table with the data and footnotes
    format_excel_table(ws, table_data, title, footnotes)
    
    # Create a BytesIO object to save the workbook
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output
